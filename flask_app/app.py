import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, flash, session
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.naive_bayes import MultinomialNB
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
import os
import re

app = Flask(__name__)
app.secret_key = 'your_secret_key'

# Paths for data files
csv_file_path = 'E:/collage/7/major project/main code/flask_app/data/expanded_keyword_dataset.csv'
excel_file_path = 'E:/collage/7/major project/main code/flask_app/data/user_data.xlsx'

# Load and preprocess the CSV file for Naive Bayes model
def load_data_and_train_model():
    # Read CSV, making sure labels are correctly interpreted
    df = pd.read_csv(csv_file_path)
    X = df['Email_Content']
    y = df['Label'].apply(lambda label: 1 if label.lower() == 'spam' else 0)  # Adjust labels for Naive Bayes
    
    # Vectorize the text data
    vectorizer = CountVectorizer()
    X_vectorized = vectorizer.fit_transform(X)

    # Train a Naive Bayes classifier
    model = MultinomialNB()
    model.fit(X_vectorized, y)

    return model, vectorizer

model, vectorizer = load_data_and_train_model()

# Helper function for Excel file operations
def load_excel_data():
    if not os.path.exists(excel_file_path):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(['Username', 'Email', 'Password'])
        wb.save(excel_file_path)

    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb.active
    return sheet, wb

# Password validation function
def validate_password(password):
    # At least 8 characters and no more than 128 characters
    if len(password) < 8 or len(password) > 128:
        return False

    # At least one uppercase letter
    if not re.search(r'[A-Z]', password):
        return False

    # At least one lowercase letter
    if not re.search(r'[a-z]', password):
        return False

    # At least one numeral
    if not re.search(r'\d', password):
        return False

    # At least one special character (~!?#@$%^&*_-+()[]{}><\/|".:;)
    if not re.search(r'[~!?#@$%^&*_\-+\(\)\[\]\{\}>/\\|".:;]', password):
        return False

    # No spaces allowed
    if re.search(r'\s', password):
        return False

    return True

# Routes for signup and login
@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        confirm_password = request.form['confirm_password']

        # Check if passwords match
        if password != confirm_password:
            flash('Passwords do not match. Please try again.', 'error')
            return redirect(url_for('signup'))

        # Validate password
        if not validate_password(password):
            flash('Password does not meet the required conditions. Please try again.', 'error')
            return redirect(url_for('signup'))

        sheet, wb = load_excel_data()

        # Check if the username already exists
        for row in sheet.iter_rows(values_only=True):
            if row[0] == username:
                flash('Username already exists. Please choose a different one.', 'error')
                return redirect(url_for('signup'))

        # Hash the password before storing it
        hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
        
        # Append new user to Excel file
        sheet.append([username, email, hashed_password])
        wb.save(excel_file_path)

        flash('You have successfully signed up! Please login.', 'success')
        return redirect(url_for('login'))

    return render_template('signup.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        sheet, wb = load_excel_data()

        # Verify username and password
        for row in sheet.iter_rows(values_only=True):
            if len(row) >= 3:  # Ensure there are at least 3 values in the row (username, email, password)
                if row[0] == username and check_password_hash(row[2], password):  # password hash is in the 3rd column
                    session['username'] = username
                    flash('Login successful!', 'success')
                    return redirect(url_for('spam_detection'))

        flash('Invalid username or password. Please try again.', 'error')
        return redirect(url_for('login'))

    return render_template('login.html')

# Route for spam detection
@app.route('/spam_detection', methods=['GET', 'POST'])
def spam_detection():
    if 'username' not in session:
        flash('Please log in to access the spam detection page.', 'error')
        return redirect(url_for('login'))

    if request.method == 'POST':
        email_content = request.form['email_content']

        if not email_content.strip():
            flash('Please enter the content of the email to check.', 'error')
            return redirect(url_for('spam_detection'))

        # Predict spam or not spam
        email_vectorized = vectorizer.transform([email_content])
        prediction = model.predict(email_vectorized)
        result = "This email is classified as Spam." if prediction[0] == 1 else "This email is classified as Not Spam."
        
        return render_template('spam_result.html', result=result)

    return render_template('spam_detection.html')

# Spam result page
@app.route('/spam_result')
def spam_result():
    return render_template('spam_result.html')

# Home route
@app.route('/')
def home():
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(debug=True)
